import drawpyo
import openpyxl


def draw_flow(page, base_x, base_y):

    document_obj = drawpyo.diagram.object_from_library(
        page=page,
        library="general",
        obj_name="document",
        value="csv",
        )

    document_obj.position = (base_x + 0, base_y + 0)

    process_obj = drawpyo.diagram.object_from_library(
        page=page,
        library="general",
        obj_name="process",
        value="New Process",
        )
    process_obj.position = (base_x + 200, base_y + 0)

    link = drawpyo.diagram.Edge(
        page=page,
        source=document_obj,
        target=process_obj,
    )

    db_obj = drawpyo.diagram.object_from_library(
        page=page,
        library="general",
        obj_name="rounded_rectangle",
        value="db",
        )
    db_obj.position = (base_x + 400, base_y + 0)

    link2 = drawpyo.diagram.Edge(
        page=page,
        source=process_obj,
        target=db_obj,
    )

def main():

    # Excel読み込み
    wb = openpyxl.load_workbook('./book.xlsx', read_only=True)
    ws = wb['Sheet1']

    for i in range(1, ws.max_row):
        print(ws.cell(row=i,column=2).value)

    # drawioファイルの書き込み
    file = drawpyo.File()
    file.file_path = r"./"
    file.file_name = "test.drawio"
    page = drawpyo.Page(file=file)

    draw_flow(page, 100, 100)
    draw_flow(page, 100, 300)

    file.write()

if __name__ == '__main__':
    main()

